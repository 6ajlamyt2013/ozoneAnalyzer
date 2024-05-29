import openpyxl
from openpyxl import Workbook

# Открываем существующий файл Excel
wb = openpyxl.load_workbook('/Users/nikolajtukmaceva/Downloads/Фильтр_по_параметрам.xlsx')
ws = wb.active

# Создаем новый рабочий лист для результата
wb_result = Workbook()
ws_result = wb_result.active
category_counts = {}

def process_excel_cell(cell_value):
    # Проверяем, является ли значение пустым
    if not cell_value:
        return ""
    
    # Проверяем, является ли значение пустой строкой
    if isinstance(cell_value, str) and cell_value.strip() == "":
        return ""
    
    # Проверяем, является ли значение числовым
    if isinstance(cell_value, (int, float)):
        return str(cell_value).lower()
    
    # Если значение не пустое, не пустая строка и не число, возвращаем его в нижнем регистре
    return str(cell_value).lower()

for row in ws.iter_rows(values_only=True):
    # Получение значений категорий
    root_category = process_excel_cell(row[0])
    category_2 = process_excel_cell(row[1])
    category_3 = process_excel_cell(row[2])
    category_4 = process_excel_cell(row[3])
    
    # Формирование ключа для словаря с использованием всех 4-х уровней категорий
    category_key = f"{root_category}_{category_2}_{category_3}_{category_4}"
    
    # Обновление счетчиков для каждой категории
    if category_key not in category_counts:
        category_counts[category_key] = 1
    else:
        category_counts[category_key] += 1

# Сохраняем результаты подсчета в новый файл
for category, count in category_counts.items():
    root_category, category_2, category_3, category_4 = category.split('_')
    ws_result.append([root_category, category_2, category_3, category_4, count])

# Сохранение результатов в файл
wb_result.save('/Users/nikolajtukmaceva/Downloads/Количество_товаров_в_категориях_4го_уровня.xlsx')
print(f"Результаты сохранены в файл: /Users/nikolajtukmaceva/Downloads/Количество_товаров_в_категориях_4го_уровня.xlsx")