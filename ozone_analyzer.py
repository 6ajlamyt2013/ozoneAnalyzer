import os
import warnings
import openpyxl
import utils
from tqdm import tqdm
from openpyxl import Workbook

class OzoneAnalyzer:
    def __init__(self, input_file, output_filtered="Фильтр_по_параметрам.xlsx", output_stats="Статистика.xlsx"):
        self.dir_path = os.path.dirname(os.path.realpath(__file__))
        self.input_file = input_file
        self.filter_output_file_path = os.path.join(self.dir_path, output_filtered)
        self.statistics_output_file_path = os.path.join(self.dir_path, output_stats)
        self.categories_to_remove = [
            'пила цепная', 'бордюр, ограждение садовые', 'газонокосилка, триммер электрические',
            'корм консервированный для кошек', 'средство увлажняющее для лица', 'шланг садовый',
            'вентилятор напольный', 'кресло складное туристическое', 'система полива',
            'дождеватели, оросители, распылители', 'жаровня, коптильня', 'секатор, сучкорез ручной',
            'грядка, клумба сборные', 'наполнитель для кошачьих туалетов', 'семена',
            'жидкий шампунь для волос женский', 'сетка антимоскитная и фурнитура',
            'стул складной туристический', 'укрывной материал для садовых растений', 'мангал, барбекю',
            'самокат', 'опора для садовых растений', 'стол складной для пикника и кемпинга',
            'удобрение (не агрохимикаты)', 'культиватор, измельчитель механические садовые',
            'горшок, кашпо', 'надувная мебель туристическая', 'солнцезащитные очки унисекс',
            'светильник садово-парковый', 'удилище, спиннинг', 'велофонарь',
            'гель жидкое средство для стирки универсальные', 'сумка/кофр для велосипеда', 'полив',
            'катушка для летней рыбалки', 'Одежда', 'подгузники детские', 'автохимия - масло моторное',
            'средство для посудомоечных машин', 'Корм сухой для собак', 'Корм сухой для кошек',
            'Ноутбуки и компьютеры', 'Обувь', 'Детское питание', 'мебель', 'смартфоны и планшеты',
            'аптека', 'крупная бытовая техника', 'детская одежда', 'напитки',
            'бакалея и кондитерские изделия'
        ]
        warnings.filterwarnings('ignore')

    def filter_products(self):
        removed_stats = {
            'scheme': 0, 'availability': 0, 'avg_orders': 0, 'price': 0,
            'catalog_views': 0, 'product_views': 0, 'lost_profit': 0
        }

        print(f"Открываем файл, пожалуйста подождите...")
        wb = openpyxl.load_workbook(self.input_file)
        ws = wb.active
        total_rows = ws.max_row
        progress_bar = tqdm(total=total_rows, desc="Фильтруем товары: ")

        wb_result = Workbook()
        ws_result = wb_result.active
        ws_result.append([str(i) for i in range(25)])  # Заголовки столбцов
        ws_result.append([
            'Категория 1 уровня', 'Категория 2 уровня', 'Категория 3 уровня', 'Категория 4 уровня',
            'Наименование товара', 'Ссылка на товар', 'Super-товар', 'Продавец', 'Бренд', 'Схема работы',
            'Сумма заказов (₽)', 'Средняя цена реализации (₽)', 'Доступность (%)',
            'Средняя сумма заказов в дни доступности (₽)', 'Среднее количество заказов в дни доступности (шт.)',
            'Сумма упущенных заказов из-за отсутствия товара (₽)', 'Количество складов отгрузки (шт.)',
            'Срок доставки (дни)', 'Количество просмотров в каталоге за 28 дней(шт.)',
            'Количество просмотров карточки товара за 28 дней (шт.)', 'Конверсия из каталога в корзину (%)',
            'Конверсия из карточки товара в корзину (%)', 'Доля рекламных расходов (%)',
            'Дата создания карточки товара'
        ])


        for row in ws.iter_rows(values_only=True):
            progress_bar.update(1)
            categories = [utils.process_excel_cell(row[i]) for i in range(4)]
            average_selling_price = utils.excel_cell_to_int(row[11])
            lost_profit = utils.excel_cell_to_int(row[15])
            availability = utils.excel_cell_to_int(row[12])
            scheme = utils.process_excel_cell(row[9])
            average_number_of_orders_per_day = utils.excel_cell_to_int(row[14])
            catalog_views = utils.excel_cell_to_int(row[18])
            product_views = utils.excel_cell_to_int(row[19])

            if any(cat.lower() in [r.lower() for r in self.categories_to_remove] for cat in categories):
                continue

            removed_stats['scheme'] += 1
            if scheme != 'fbo':
                continue

            if availability >= 0.8:
                removed_stats['availability'] += 1
                continue

            if average_number_of_orders_per_day <= 5:
                removed_stats['avg_orders'] += 1
                continue

            if not (400 < average_selling_price < 5500):
                removed_stats['price'] += 1
                continue

            if catalog_views <= 20000:
                removed_stats['catalog_views'] += 1
                continue

            if product_views <= 2000:
                removed_stats['product_views'] += 1
                continue

            if lost_profit <= 100000:
                removed_stats['lost_profit'] += 1
                continue

            ws_result.append(row)

        wb_result.save(self.filter_output_file_path)
        progress_bar.close()

        print("\nСтатистика удаления товаров:")
        for key, value in removed_stats.items():
            print(f"Удалено по {key}: {value}")
        print(f"Результат сохранен в файл: {self.filter_output_file_path}")


    def generate_statistics(self):
        print(f"Подготавливаем статистику по категориям")
        wb = openpyxl.load_workbook(self.filter_output_file_path)
        ws = wb.active

        wb_result = Workbook()
        ws_result = wb_result.active
        ws_result.append(['Категория 1', 'Категория 2', 'Категория 3', 'Категория 4', 'Количество товаров', 'Оборот', 'Продавцов', 'Количество товаров на одного продавца', 'Количество продаж в день'])

        category_counts = {}

        for row in ws.iter_rows(values_only=True):
            avg_orders = utils.excel_cell_to_int(row[14])
            categories = [utils.process_excel_cell(row[i]) for i in range(4)]
            turnover = utils.excel_cell_to_int(row[10])
            seller = utils.process_excel_cell(row[7])

            category_key = "_".join(categories)

            if category_key not in category_counts:
                category_counts[category_key] = {
                    'count': 0, 'turnover': 0, 'sellers': set(), 'avg_orders': 0
                }

            category_counts[category_key]['count'] += 1
            category_counts[category_key]['turnover'] += turnover
            category_counts[category_key]['sellers'].add(seller)
            category_counts[category_key]['avg_orders'] += avg_orders

        for category, data in category_counts.items():
            categories = category.split("_")
            sellers_count = len(data['sellers'])
            avg_items_per_seller = data['count'] / sellers_count if sellers_count else 0
            if avg_items_per_seller > 4:
               ws_result.append(categories + [data['count'], data['turnover'], sellers_count, avg_items_per_seller, data['avg_orders']])


        wb_result.save(self.statistics_output_file_path)
        print(f"Статистика сгенерирована, сохранена: {self.statistics_output_file_path}")
