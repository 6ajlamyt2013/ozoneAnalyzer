import re
import os
import time
import openpyxl
import warnings
from tqdm import tqdm
from openpyxl import Workbook

dir_path = os.path.dirname(os.path.realpath(__file__))
warnings.filterwarnings('ignore')

all_goods = input("Укажите путь к файлу с товарами: ")
filter_output_file_path = os.path.join(dir_path, "Фильтр_по_параметрам.xlsx")
statistics_output_file_path = os.path.join(dir_path, "Статистика.xlsx")

# Категории для удаления
categories_to_remove = ['пила цепная', 
'бордюр, ограждение садовые', 
'газонокосилка, триммер электрические', 
'корм консервированный для кошек', 
'средство увлажняющее для лица', 
'шланг садовый', 'вентилятор напольный', 
'кресло складное туристическое', 'система полива', 
'дождеватели, оросители, распылители', 'жаровня, коптильня', 
'секатор, сучкорез ручной', 'грядка, клумба сборные', 
'наполнитель для кошачьих туалетов', 'семена', 
'жидкий шампунь для волос женский', 
'сетка антимоскитная и фурнитура', 
'стул складной туристический', 
'укрывной материал для садовых растений', 
'мангал, барбекю', 'самокат', 
'опора для садовых растений', 
'стол складной для пикника и кемпинга', 
'удобрение (не агрохимикаты)', 
'культиватор, измельчитель механические садовые', 
'горшок, кашпо', 'надувная мебель туристическая', 
'солнцезащитные очки унисекс', 
'светильник садово-парковый', 
'удилище, спиннинг', 'велофонарь', 
'гель жидкое средство для стирки универсальные', 
'сумка/кофр для велосипеда', 'полив',
'катушка для летней рыбалки', 'Одежда', 
'подгузники детские', 
'автохимия - масло моторное', 
'средство для посудомоечных машин', 
'Корм сухой для собак', 
'Корм сухой для кошек', 
'Ноутбуки и компьютеры', 
'Обувь', 'Детское питание', 'мебель', 
'смартфоны и планшеты', 'аптека', 
'крупная бытовая техника', 'детская одежда', 'напитки', 
'бакалея и кондитерские изделия']

#Метод берет любое значение ячейки и возвращает int 
def excel_cell_to_int(cell_value):
    # Проверяем, не пустая ли ячейка
    if cell_value is None:
        return 0
    
    # Преобразуем значение в строку, если оно не строка
    if not isinstance(cell_value, str):
        cell_value = str(cell_value)
    
    # Удаляем все нецифровые символы, кроме запятых и точек (для чисел с плавающей точкой)
    cell_value = re.sub(r'[^\d.,]', '', cell_value)
    cell_value = cell_value.replace(" ","")

    # Если после удаления спецсимволов ничего не осталось или осталась только точка, возвращаем 0
    if not cell_value or cell_value == '.':
        return 0
    
    # Если значение содержит запятые или точки, то это может быть число с плавающей точкой
    if ',' in cell_value or '.' in cell_value:
        # Заменяем запятые на точки, чтобы не потерять дробную часть
        cell_value = cell_value.replace(',', '.')
        # Преобразуем в float и округляем до ближайшего целого
        return int(float(cell_value))
    else:
        # Преобразуем в int
        return int(cell_value)

#Метод берет любое значение ячейки и возвращает str
def process_excel_cell(cell_value):
    # Проверяем, является ли значение пустым
    if not cell_value:
        return ""
    
    # Проверяем, является ли значение пустой строкой
    if isinstance(cell_value, str) and cell_value.strip() == "":
        return ""
    
    # Проверяем, является ли значение числовым
    if isinstance(cell_value, (int, float)):
        return str(cell_value).lower()
    
    # Если значение не пустое, не пустая строка и не число, возвращаем его в нижнем регистре
    return str(cell_value).lower()

#Метод принимает на вход путь к исходной таблице и фильтрует по заданным параметрам
def filter_producs_by_parametrs(path):
    # Счетчики удаленных строк
    removed_by_scheme = 0
    removed_by_availability = 0
    removed_by_avg_orders = 0
    removed_by_price = 0
    removed_by_catalog_views = 0
    removed_by_product_views = 0
    removed_by_lost_profit = 0

    # Открываем существующий файл Excel
    print(f"Открываем файл, пожалуйста подождите...")
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Получите общее количество строк
    total_rows = ws.max_row

    # Создайте прогресс бар
    progress_bar = tqdm(total=total_rows, desc="Фильтруем товары: ")
    
    # Создаем новый рабочий лист для результата
    wb_result = Workbook()
    ws_result = wb_result.active
    ws_result.append([str(i) for i in range(25)])
    ws_result.append(['Категория 1 уровня', 'Категория 2 уровня', 'Категория 3 уровня', 'Категория 4 уровня', 'Наименование товара', 'Ссылка на товар', 'Super-товар', 'Продавец', 'Бренд', 'Схема работы', 'Сумма заказов (₽)', 'Средняя цена реализации (₽)', 'Доступность (%)', 'Средняя сумма заказов в дни доступности (₽)', 'Среднее количество заказов в дни доступности (шт.)', 'Сумма упущенных заказов из-за отсутствия товара (₽)', 'Количество складов отгрузки (шт.)', 'Срок доставки (дни)', 'Количество просмотров в каталоге за 28 дней(шт.)', 'Количество просмотров карточки товара за 28 дней (шт.)', 'Конверсия из каталога в корзину (%)', 'Конверсия из карточки товара в корзину (%)', 'Доля рекламных расходов (%)', 'Дата создания карточки товара'])
    
    # Проходимся по каждой строке данных
    for row in ws.iter_rows(values_only=True):

        progress_bar.update(1)

        # Получение значения из таблицы 
        root_category = process_excel_cell(row[0])
        category_2_lvl = process_excel_cell(row[1])
        category_3_lvl = process_excel_cell(row[2])
        category_4_lvl = process_excel_cell(row[3])
        categories = [root_category, category_2_lvl, category_3_lvl, category_4_lvl]
    
        average_selling_price = excel_cell_to_int(row[11])
        lost_profit = excel_cell_to_int(row[15])
        availability = excel_cell_to_int(row[12])
        sheme = process_excel_cell(row[9])
        average_number_of_orders_per_day = excel_cell_to_int(row[14])
        number_of_catalog_views_in_28_days= excel_cell_to_int(row[18])
        number_of_product_card_views_in_28_days= excel_cell_to_int(row[19])

        # Проверка категорий 
        match_found = False
        for category in categories:
            if category.lower() in [cat.lower() for cat in categories_to_remove]:
                match_found = True
                break
    
        if match_found:
            continue

        removed_by_scheme += 1
        if not(sheme != 'FBO'):
            continue

        if not(availability < 0.8):
            removed_by_availability += 1
            continue

        if not(average_number_of_orders_per_day > 5):
            removed_by_avg_orders += 1
            continue
    
        # Проверка стоимости товара
        if not(average_selling_price > 400 and average_selling_price < 5500):
            removed_by_price += 1
            continue  
    
        # Проверка просмотров каталога за 28 дней
        if not(number_of_catalog_views_in_28_days > 20000):
            removed_by_catalog_views += 1
            continue
        
        # Проверка просмотров карточки товара за 28 дней
        if not(number_of_product_card_views_in_28_days > 2000):
            removed_by_product_views += 1
            continue

        if not(lost_profit > 100000):
            removed_by_lost_profit += 1
            continue

        # Если товар прошел все проверки, добавляем его в новый список
        ws_result.append(row)
    
        # Форматируем вывод    
        #print(f"Категория: {root_category:<20} | Стоимость: {average_selling_price:>7} | Продажи в день: {average_number_of_orders_per_day:>7} | Просмотры каталога: {number_of_catalog_views_in_28_days:>7} | Просмотры карточки: {number_of_product_card_views_in_28_days:>7}")
        
    # Сохраняем результат в новый файл Excel
    wb_result.save(filter_output_file_path)
    progress_bar.close()

     # Форматируем и выводим результат
    print(f"""
    Статистика удаления товаров:
    ----------------------------
    Удалено по схеме работы:       {removed_by_scheme}
    Удалено по доступности:        {removed_by_availability}
    Удалено по средним заказам:    {removed_by_avg_orders}
    Удалено по цене:               {removed_by_price}
    Удалено по просмотрам каталога:{removed_by_catalog_views}
    Удалено по просмотрам товара:  {removed_by_product_views}
    Удалено по упущенной прибыли:  {removed_by_lost_profit}
    """)
    
    print(f"Результат сохранен в файл: {filter_output_file_path}")

# Генерация статистики по категориям
def statistics_generator(excel_file_path):

    print(f"Подготавливаем статистику по категориям")

    # Открытие существующего файла Excel
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active

    # Создание нового рабочего листа для результата
    wb_result = Workbook()
    ws_result = wb_result.active
    ws_result.append(['Категория 1', 'Категория 2', 'Категория 3', 'Категория 4', 'Количество товаров', 'Оборот', 'Продавцов', 'Количество товаров на одного продавца', 'Количество продаж в день'])

    category_counts = {}

    for row in ws.iter_rows(values_only=True):
        average_number_orders_per_days_availability = excel_cell_to_int(row[14])
        root_category = process_excel_cell(row[0])
        category_2 = process_excel_cell(row[1])
        category_3 = process_excel_cell(row[2])
        category_4 = process_excel_cell(row[3])
        turnover = excel_cell_to_int(row[10])
        seller = process_excel_cell(row[7])

        category_key = f"{root_category}_{category_2}_{category_3}_{category_4}"

        if category_key not in category_counts:
            category_counts[category_key] = {
                'count': 0,
                'turnover': 0,
                'sellers': set(),
                'average_number_orders_per_days_availability': 0
            }

        category_counts[category_key]['count'] += 1
        category_counts[category_key]['turnover'] += turnover
        category_counts[category_key]['sellers'].add(seller)
        category_counts[category_key]['average_number_orders_per_days_availability'] += average_number_orders_per_days_availability

    for category, data in category_counts.items():
        root_category, category_2, category_3, category_4 = category.split('_')
        sellers_count = len(data['sellers'])
        average_items_per_seller = data['count'] / sellers_count if sellers_count else 0  # Избегаем деления на ноль
        if average_items_per_seller > 4:
            ws_result.append([
                root_category,
                category_2,
                category_3,
                category_4,
                data['count'],
                data['turnover'],
                sellers_count,
                average_items_per_seller,
                data['average_number_orders_per_days_availability']
            ])
    
    wb_result.save(statistics_output_file_path)
    print(f"Статистика сгенерирована, сохранена: {statistics_output_file_path}")

filter_producs_by_parametrs(all_goods)
statistics_generator(filter_output_file_path)
